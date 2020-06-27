import openpyxl


def getData(sheet):
    """
    Extracts data from origin file, separating between students’ information
    and exam answers. Returns it as tuples: one tuple per student, containing
    a tuple with their information and another tuple with their answers.
    Inputs: sheet, an Excel sheet object.
    Returns: a tuple.
    """
    info, answers = [], []
    for rowNum in range(2, sheet.max_row + 1):
        infoTemp, answersTemp = [], []
        for colNum in range(2, 6):
            infoTemp.append(sheet.cell(row=rowNum, column=colNum).value)
        for i in range(4, 6):
            infoTemp.insert(4, infoTemp[0])
            del infoTemp[0]
        info.append(tuple(infoTemp))
        for colNum in range(7, 78):
            content = sheet.cell(row=rowNum, column=colNum).value
            # Countermeasure for error in exam. Remove block if exam is fixed.
            try:
                answersTemp.append(content[0] if "," not in content else None)
            except TypeError:
                answersTemp.append(None)
        answers.append(tuple(answersTemp))
    return tuple(info), tuple(answers)


def getAnswerKey():
    """
    Extract answer key from file.
    Returns: a tuple of strings.
    """
    wb = openpyxl.load_workbook("answerKey.xlsx")
    sheet = wb.active
    return tuple([
        sheet.cell(row=row, column=3).value
        for row in range(1, sheet.max_row + 1)
    ])


def checkAnswers(answers, answerKey):
    """
    Checks answers.
    Inputs answers: a list of lists of strings.
    Returns: a tuple of tuples of integers.
    """
    def checkSection(current, start, end):
        """
        Check answers in a section.
        Inputs: current, a list of strings; start, an integer; end, an integer.
        Returns: an integer.
        """
        counter = 0
        for i in range(start, end):
            if current[i] == answerKey[i]:
                counter += 1
        return counter

    final = []
    for elem in answers:
        results = [
            checkSection(elem, 0, 12),
            checkSection(elem, 12, 24),
            checkSection(elem, 24, 36),
            checkSection(elem, 36, 47),
            checkSection(elem, 47, 57),
            checkSection(elem, 57, 71)
        ]
        final.append(tuple(results))
    return tuple(final)


def writeResults():
    """
    Write results to a new file.
    Returns: nothing.
    """
    wb = openpyxl.load_workbook("marks.xlsx")
    sheet = wb.active
    info = getBasics(sheet)
    answers = getAnswers(sheet)
    answerKey = getAnswerKey()
    checkedAnswers = checkAnswers(answers, answerKey)

    wb = openpyxl.Workbook()
    sheet = wb.active
    values = ("First Name", "Last Name", "Email", "Section 1", "Section 2",
              "Section 3", "Section 4", "Section 5", "Section 6", "Old score",
              "New score")
    for column in range(1, len(values) + 1):
        sheet.cell(row=1, column=column).value = values[column - 1]
    for row in range(len(info)):
        sheet.cell(row=row + 2, column=1).value = info[row][2]
        sheet.cell(row=row + 2, column=2).value = info[row][3]
        sheet.cell(row=row + 2, column=3).value = info[row][0]
        sheet.cell(row=row + 2, column=10).value = info[row][1]
    for row in range(len(checkedAnswers)):
        for column in range(4, 10):
            sheet.cell(
                row=row + 2, column=column
            ).value = checkedAnswers[row][column - 4]
        sheet.cell(row=row + 2, column=11).value = sum(checkedAnswers[row])
    wb.save("results.xlsx")


writeResults()
