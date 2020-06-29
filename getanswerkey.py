"""Scrape Excel file for answer key and saves it in another file."""

import openpyxl


def get_answer_key():
    """
    Extract answer key from file.

    Returns: a tuple of strings.
    """
    wb = openpyxl.load_workbook("answerKey.xlsx")
    sheet = wb.active
    return tuple([
        sheet.cell(row=row, column=3).value
        for row in range(1, sheet.max_row + 1)
    ])


def write_answer_key():
    """Write answer key to file."""
    file = open("answerkey.py", "w")
    file.write(f"answer_key = {get_answer_key()}")
    file.close()


write_answer_key()
