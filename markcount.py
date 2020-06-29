"""
Mark exam sections for each student and saves results in a file.

Students’ information is parsed from the original file. Their answers
are checked against the stored answers key, return the marks for each
exam section. Information and results are then saved in results.xlsx.
"""

import openpyxl

from answerkey import answer_key


def get_data(sheet):
    """
    Parse information from origin file.

    Inputs: sheet, an Excel sheet object.
    Returns: tuples of various data types.
    """
    info, answers = [], []
    for rownum in range(2, sheet.max_row + 1):
        temp1, temp2 = [], []
        for colnum in range(2, 6):
            temp1.append(sheet.cell(row=rownum, column=colnum).value)
        for i in range(4, 6):
            temp1.insert(4, temp1[0])
            del temp1[0]
        info.append(tuple(temp1))
        for colnum in range(7, 78):
            content = sheet.cell(row=rownum, column=colnum).value
            # Countermeasure for error in exam. Remove block if fixed.
            try:
                temp2.append(content[0] if "," not in content else None)
            except TypeError:
                temp2.append(None)
        answers.append(tuple(temp2))
    return tuple(info), tuple(answers)


def check_answers(answers, answer_key):
    """
    Check students’ answers against answer key.

    Inputs: answers, a tuple of tuples of strs.
            answer_key, a tuple of strs.
    Returns: tuple of tuples of ints.
    """

    def check_section(current, start, end):
        """
        Mark answers in a section.

        Inputs: current, a list of strs.
                start, an int.
                end, an int.
        Returns: an int.
        """
        counter = 0
        for i in range(start, end):
            if current[i] == answer_key[i]:
                counter += 1
        return counter

    final = []
    for elem in answers:
        results = [
            check_section(elem, 0, 12),
            check_section(elem, 12, 24),
            check_section(elem, 24, 36),
            check_section(elem, 36, 47),
            check_section(elem, 47, 57),
            check_section(elem, 57, 71),
        ]
        final.append(tuple(results))
    return tuple(final)


def write_results():
    """Write collected information to a new Excel file."""
    wb = openpyxl.load_workbook("marks.xlsx")
    sheet = wb.active
    info, answers = get_data(sheet)
    sections = check_answers(answers, answer_key)

    wb = openpyxl.Workbook()
    sheet = wb.active
    values = ("First Name", "Last Name", "Email", "Section 1", "Section 2",
              "Section 3", "Section 4", "Section 5", "Section 6", "Old score",
              "New score")
    for colnum in range(len(values)):
        sheet.cell(row=1, column=colnum + 1).value = values[colnum]
    for rownum in range(2, len(info) + 2):
        for colnum in range(1, 4):
            sheet.cell(
                row=rownum,
                column=colnum,
            ).value = info[rownum - 2][colnum - 1]
        for colnum in range(4, 10):
            sheet.cell(
                row=rownum,
                column=colnum,
            ).value = sections[rownum - 2][colnum - 4]
        sheet.cell(row=rownum, column=10).value = info[rownum - 2][3]
        sheet.cell(row=rownum, column=11).value = sum(sections[rownum - 2])
    wb.save("results.xlsx")


write_results()
