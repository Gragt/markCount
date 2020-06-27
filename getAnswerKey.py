import openpyxl


def getAnswerKey():
    """
    Extract answer key from file.
    Returns: a tuple of strings.
    """
    wb = openpyxl.load_workbook("answerKey.xlsx")
    sheet = wb.active
    return tuple([
        sheet.cell(row=row, column=3).value
        for row in range(1, sheet.max_row + 1)
    ])


def writeAnswerKey():
    """
    Write answer key to a file.
    Returns: nothing
    """
    file = open("answerKey.py", "w")
    file.write(f"answerKey = {getAnswerKey()}")
    file.close()


writeAnswerKey()
